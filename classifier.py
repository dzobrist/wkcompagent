"""
Payroll extraction, mapping-rule loading, and exclusion arithmetic.
Claude handles classification decisions (see processor.py); everything
here is deterministic Python.
"""

import re
from typing import Optional
import openpyxl

CLASS_CODE_ORDER = [
    "CO8810", "CO8869", "CO9180",
    "FL8742", "FL9180",
    "GA9180",
    "VA8742", "VA9180",
    "WV8742", "WV8810", "WV9180",
]

CO9180_SUBGROUP_DEPTS = {
    "Soaring Guides — Depts 15 & 20": {15, 20},
    "Outfitter Guides — Depts 9 & 21": {9, 21},
    "Falconry - Broadmoor — Dept 18": {18},
}

# Depts 16 & 17 are classified as CO8810 on the carrier form but are tracked
# separately here so the internal PDF can show them alongside the Soaring CO9180
# rows for cost-allocation visibility.
SOARING_CO8810_DEPTS = {16, 17}


def _dept_num(dept: str) -> Optional[int]:
    m = re.match(r'^\s*(\d+)', dept)
    try:
        return int(m.group(1)) if m else None
    except (ValueError, AttributeError):
        return None


def _find_header_row(ws, marker: str, col: int = 0) -> int:
    """Return the 1-based row number of the row where ws.cell(row, col+1) == marker.
    Returns 0 if not found."""
    for row in ws.iter_rows(values_only=False):
        cell = row[col]
        if cell.value and str(cell.value).strip() == marker:
            return cell.row
    return 0


def load_mapping_rules(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)

    # ── Department → Class Code ──────────────────────────────────────────────
    # Sheet name: "Department to Class Code"
    # Columns (after header row): col0=Department, col1=State, col2=Class Code
    dept_to_code: dict[str, str] = {}
    ws_dept = wb["Department to Class Code"]
    hdr = _find_header_row(ws_dept, "Department (exact text from source)", col=0)
    if hdr == 0:
        hdr = 1  # fallback: start from row 1
    for row in ws_dept.iter_rows(min_row=hdr + 1, values_only=True):
        dept = row[0]
        code = row[2] if len(row) > 2 else None
        if dept and code:
            dept_str = str(dept).strip()
            code_str = str(code).strip()
            # Skip trailing-note rows (dept cell contains long instructional text)
            if dept_str and code_str and len(dept_str) < 120:
                dept_to_code[dept_str] = code_str

    # ── Earning Type Rules ───────────────────────────────────────────────────
    # Sheet name: "Earning Type Rules"
    # Columns (after header row): col0=Earning Type, col1=Gross Treatment, col2=Excluded Treatment
    earning_rules: dict[str, str] = {}
    ws_earn = wb["Earning Type Rules"]
    hdr = _find_header_row(ws_earn, "Earning Type (exact text)", col=0)
    if hdr == 0:
        # Fall back: search headers for "excl" keyword
        headers = [str(c).strip().lower() if c else ""
                   for c in next(ws_earn.iter_rows(min_row=1, max_row=1, values_only=True))]
        try:
            excl_col = next(i for i, h in enumerate(headers) if "excl" in h)
        except StopIteration:
            excl_col = 2
        data_start = 2
    else:
        excl_col = 2
        data_start = hdr + 1

    for row in ws_earn.iter_rows(min_row=data_start, values_only=True):
        if not row[0]:
            continue
        earn_type = str(row[0]).strip()
        # Skip long instructional note rows
        if len(earn_type) > 80:
            continue
        excl_val = str(row[excl_col]).strip().lower() if len(row) > excl_col and row[excl_col] is not None else "none"
        earning_rules[earn_type] = excl_val

    # ── Employee Overrides ───────────────────────────────────────────────────
    # Sheet name: "Employee Overrides"
    # Columns (after header row): col0=Name, col1=Dept, col2=Class Code, col3=Action, col4=Reason
    overrides: list[dict] = []
    ws_ov = wb["Employee Overrides"]
    hdr = _find_header_row(ws_ov, "Employee Name (exact text from source)", col=0)
    data_start = hdr + 1 if hdr else 2

    for row in ws_ov.iter_rows(min_row=data_start, values_only=True):
        name = row[0] if row[0] else None
        action = row[3] if len(row) > 3 else None
        # Only accept rows with a valid Action value
        if not name or not action:
            continue
        action_str = str(action).strip()
        if action_str not in ("Reassign", "Exclude"):
            continue
        # Skip long instructional note rows masquerading as data
        if len(str(name).strip()) > 120:
            continue
        dept_raw = str(row[1]).strip() if len(row) > 1 and row[1] else None
        if dept_raw and re.match(r'^\[.*\]$|^n/?a$|^tbd$|^placeholder$', dept_raw, re.IGNORECASE):
            dept_raw = None
        overrides.append({
            "name": str(name).strip(),
            "dept": dept_raw if dept_raw else None,
            "class_code": str(row[2]).strip() if len(row) > 2 and row[2] else None,
            "action": action_str,
            "reason": str(row[4]).strip() if len(row) > 4 and row[4] else "",
        })

    return {
        "dept_to_code": dept_to_code,
        "earning_rules": earning_rules,
        "overrides": overrides,
    }


def _excluded_amount(amount: float, mode: str) -> float:
    mode = mode.lower().strip()
    if not mode or mode in ("none", "0", "zero", "$0", "no", "false"):
        return 0.0
    if mode in ("full", "all", "yes", "100%", "100", "true"):
        return amount
    if "third" in mode or "1/3" in mode or "33" in mode or "premium" in mode:
        return round(amount / 3, 2)
    try:
        pct = float(mode.strip("%")) / 100
        return round(amount * pct, 2)
    except ValueError:
        return 0.0


def extract_payroll_rows(ws) -> list[dict]:
    rows = []
    consecutive_blanks = 0
    for row_idx in range(8, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        if val is None:
            consecutive_blanks += 1
            if consecutive_blanks >= 2:
                break
            continue
        name = str(val).strip()
        if name in ("Company Totals", "Department Totals"):
            break
        consecutive_blanks = 0
        dept = str(ws.cell(row=row_idx, column=2).value or "").strip()
        earnings = []
        for slot in range(9):  # 9 earning slots: C-F through AI-AL
            col = 3 + slot * 4
            earn_type = ws.cell(row=row_idx, column=col).value
            if earn_type:
                amount = ws.cell(row=row_idx, column=col + 3).value
                if amount is not None:
                    try:
                        earnings.append({"type": str(earn_type).strip(), "amount": float(amount)})
                    except (ValueError, TypeError):
                        pass
        rows.append({"row": row_idx, "employee": name, "department": dept, "earnings": earnings})
    return rows
