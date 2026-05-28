"""
Payroll processor: Claude classifies (class code, overrides, exceptions),
Python aggregates all arithmetic. This gives project-accurate classification
with deterministic dollar math.
"""

import io
import os
import re
from datetime import date, timedelta

import anthropic
import openpyxl

from classifier import (
    CLASS_CODE_ORDER,
    CO9180_SUBGROUP_DEPTS,
    SOARING_CO8810_DEPTS,
    _dept_num,
    _excluded_amount,
    extract_payroll_rows,
    load_mapping_rules,
)
from email_service import send_error_email, send_result_email
from report_generator import generate_broadmoor_pdf, generate_pdf

KNOWLEDGE_DIR = os.path.join(os.path.dirname(__file__), "knowledge")
MAPPING_RULES_PATH = os.path.join(KNOWLEDGE_DIR, "Mapping_Rules.xlsx")
FORM_TEMPLATE_PATH = os.path.join(KNOWLEDGE_DIR, "Form_Template.xlsx")

CODE_TO_ROW = {code: 17 + i for i, code in enumerate(CLASS_CODE_ORDER)}

# Departments whose employee-level detail appears in the CO9180 subgroup PDF section
SUBGROUP_DEPT_NUMS = {9, 15, 16, 17, 18, 20, 21}

# ── System prompt (project instructions — classification focus) ───────────────
CLASSIFICATION_SYSTEM_PROMPT = """You are a payroll classification assistant for Greenbrier Outfitters and affiliated properties.

Your ONLY task is to read the payroll rows and Mapping Rules and return classification decisions for each row via the submit_classification tool. Do NOT compute any dollar totals — Python handles all arithmetic.

The reporting carrier is Applied Underwriters. The named insured is Greenbrier Outfitters. The producer is Ashley Cazire. The policy number is 37-664725-01-01; -01-02. The policy period is 04/15/26–04/15/27.

## Source file format (Earnings by Department)

Rows 1–6: report header metadata — skip.
Row 7: column headers.
Rows 8+: employee detail rows — classify these.
Stop at two consecutive blank rows in column A, or when column A = "Company Totals" / "Department Totals".

Each detail row has up to 9 earning slots:
- Slot 1: Columns C–F (Type, Hours, Rate, Amount)
- Slot 2: Columns G–J
- Slot 3: Columns K–N
- Slot 4: Columns O–R
- Slot 5: Columns S–V
- Slot 6: Columns W–Z
- Slot 7: Columns AA–AD
- Slot 8: Columns AE–AH
- Slot 9: Columns AI–AL

Slots with a blank earning type are empty — skip those. Check all 9 slots every time.

## Mapping_Rules.xlsx structure

Tabs you actively use:
- "Department to Class Code": Department name → Class Code. Default for all employees.
- "Earning Type Rules": Earning type → gross + excluded treatment.
- "Employee Overrides": Per-employee rules that take precedence over dept lookup.
  Columns: Employee Name | Department (optional) | Class Code | Action | Reason

## Classification rules

**Step 1: Check Employee Overrides first.**
Find all rows in Employee Overrides where Employee Name matches column A of the source row exactly
(comma, middle initial, capitalization, spacing — exact match only).

Among name matches:
- Prefer a row where the override's Department also matches the source row's Department exactly.
- If no department-specific row matches, use the first row with a blank Department field.
- If two rows tie, use the first one and add a data quality note.

Once the winning override row is found:
- Action = Exclude → disposition = "exclude". Drop the entire source row.
- Action = Reassign → disposition = "include", class_code = the override's Class Code.

**Step 2: If no override matches**, look up the Department in "Department to Class Code".
- If found → disposition = "include", class_code = mapped value.
- If not found → disposition = "exception".

**Step 3: Flag exception earnings.**
For each earning slot on an "include" row, check if the earning type exists in "Earning Type Rules".
If not found, add it to exception_earnings for that row.

## Hard rules

- Name matching is exact. "Smith, John A" ≠ "Smith, John". Flag near-misses as data quality notes.
- Department matching is exact. "1 - OFFICE" ≠ "1 - Office".
- Exclude overrides drop the ENTIRE source row — all earning slots omitted.
- Never include "Company Totals" or "Department Totals" rows.
- Always check all 9 earning slots per row.
- Override precedence is absolute: Employee Overrides always beat department lookup.
- Never invent data. If unclear, record as exception.
"""

CLASSIFICATION_TOOL = {
    "name": "submit_classification",
    "description": (
        "Submit classification decisions for each payroll row. "
        "Classify disposition and class code only — do NOT compute dollar totals."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "classified_rows": {
                "type": "array",
                "description": "One entry per employee detail row",
                "items": {
                    "type": "object",
                    "properties": {
                        "row_num": {"type": "integer"},
                        "name": {"type": "string"},
                        "department": {"type": "string"},
                        "disposition": {
                            "type": "string",
                            "enum": ["include", "exclude", "exception"],
                        },
                        "class_code": {
                            "type": "string",
                            "description": "Required when disposition=include",
                        },
                        "override_applied": {"type": "boolean"},
                        "override_action": {
                            "type": "string",
                            "enum": ["Reassign", "Exclude"],
                        },
                        "override_reason": {"type": "string"},
                        "exception_reason": {
                            "type": "string",
                            "description": "Required when disposition=exception",
                        },
                        "exception_earnings": {
                            "type": "array",
                            "description": "Earning types not found in Earning Type Rules",
                            "items": {"type": "string"},
                        },
                    },
                    "required": ["row_num", "name", "department", "disposition"],
                },
            },
            "data_quality_notes": {
                "type": "array",
                "items": {"type": "string"},
            },
        },
        "required": ["classified_rows"],
    },
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _read_mapping_text() -> str:
    """Convert Mapping_Rules.xlsx to readable text for Claude."""
    wb = openpyxl.load_workbook(MAPPING_RULES_PATH, data_only=True)
    sections = []
    for sheet_name in ["Department to Class Code", "Earning Type Rules", "Employee Overrides"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows_text = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            if any(c is not None for c in row):
                rows_text.append(" | ".join(str(c) if c is not None else "" for c in row))
        sections.append(f"=== {sheet_name} ===\n" + "\n".join(rows_text))
    return "\n\n".join(sections)


def _payroll_rows_to_text(rows: list[dict]) -> str:
    """Format payroll rows as readable text for Claude."""
    lines = [f"PAYROLL DATA — {len(rows)} employee detail rows\n"]
    for r in rows:
        parts = []
        for i, e in enumerate(r["earnings"], 1):
            parts.append(f"[Slot {i}] {e['type']}: ${e['amount']:.2f}")
        earnings_str = " | ".join(parts) if parts else "(no earnings)"
        lines.append(
            f"Row {r['row']}: {r['employee']} | Dept: {r['department']} | {earnings_str}"
        )
    return "\n".join(lines)


async def _classify_with_claude(payroll_rows: list[dict], period_date: str) -> dict:
    """Call Claude for classification decisions only. Returns classified_rows dict."""
    client = anthropic.AsyncAnthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
    mapping_text = _read_mapping_text()
    payroll_text = _payroll_rows_to_text(payroll_rows)

    response = await client.messages.create(
        model=os.getenv("CLAUDE_MODEL", "claude-opus-4-5"),
        max_tokens=16000,
        system=[{
            "type": "text",
            "text": CLASSIFICATION_SYSTEM_PROMPT,
            "cache_control": {"type": "ephemeral"},
        }],
        tools=[CLASSIFICATION_TOOL],
        tool_choice={"type": "any"},
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "text",
                    "text": f"MAPPING RULES\n\n{mapping_text}",
                    "cache_control": {"type": "ephemeral"},
                },
                {
                    "type": "text",
                    "text": (
                        f"Reporting period end: {period_date}\n\n"
                        f"Classify each row below. Return classification decisions only — "
                        f"do not compute any dollar totals.\n\n{payroll_text}"
                    ),
                },
            ],
        }],
    )

    for block in response.content:
        if block.type == "tool_use" and block.name == "submit_classification":
            return block.input

    raise ValueError("Claude did not return a classification. Check API logs.")


def _aggregate(payroll_rows: list[dict], classification: dict, mapping_rules: dict) -> dict:
    """
    Apply Claude's classification decisions and compute all arithmetic in Python.
    Claude decides who goes where. Python does all the math.
    """
    earning_rules = mapping_rules["earning_rules"]
    rows_by_num = {r["row"]: r for r in payroll_rows}

    totals = {code: {"gross": 0.0, "excluded": 0.0} for code in CLASS_CODE_ORDER}
    co9180_subs = {name: {"gross": 0.0, "excluded": 0.0} for name in CO9180_SUBGROUP_DEPTS}
    soaring_co8810 = {"gross": 0.0, "excluded": 0.0}

    # Employee detail for PDF subgroups: dept -> {name -> {gross, excluded, class_code}}
    subgroup_detail: dict[str, dict] = {}

    overrides_applied = []
    exceptions = []
    dq_notes = classification.get("data_quality_notes", [])
    earning_count = 0

    for cr in classification.get("classified_rows", []):
        row_num = cr["row_num"]
        name = cr["name"]
        dept = cr["department"]
        disposition = cr["disposition"]

        if disposition == "exclude":
            reason = cr.get("override_reason", "")
            overrides_applied.append(f"{name} / {dept} / Exclude — {reason}")
            continue

        if disposition == "exception":
            reason = cr.get("exception_reason", "no class code mapping")
            exceptions.append(f"Row {row_num}: {name} / dept '{dept}' — {reason}; row excluded")
            continue

        class_code = cr.get("class_code")
        if not class_code or class_code not in totals:
            exceptions.append(f"Row {row_num}: {name} / unknown class code '{class_code}'")
            continue

        if cr.get("override_applied") and cr.get("override_action") == "Reassign":
            reason = cr.get("override_reason", "")
            overrides_applied.append(f"{name} / {dept} / Reassign → {class_code} — {reason}")

        payroll_row = rows_by_num.get(row_num)
        if not payroll_row:
            exceptions.append(f"Row {row_num}: {name} — row data not found in source")
            continue

        exception_earning_types = set(cr.get("exception_earnings") or [])
        dept_num = _dept_num(dept)

        for earning in payroll_row.get("earnings", []):
            earn_type = earning["type"]
            amount = earning["amount"]
            earning_count += 1

            if earn_type in exception_earning_types or earn_type not in earning_rules:
                exceptions.append(
                    f"Row {row_num}: {name} — unrecognized earning type '{earn_type}' "
                    f"(${amount:,.2f}); excluded from totals"
                )
                continue

            excl = _excluded_amount(amount, earning_rules[earn_type])
            totals[class_code]["gross"] = round(totals[class_code]["gross"] + amount, 2)
            totals[class_code]["excluded"] = round(totals[class_code]["excluded"] + excl, 2)

            # CO9180 subgroup tracking
            if class_code == "CO9180":
                for sg_name, nums in CO9180_SUBGROUP_DEPTS.items():
                    if dept_num in nums:
                        co9180_subs[sg_name]["gross"] = round(co9180_subs[sg_name]["gross"] + amount, 2)
                        co9180_subs[sg_name]["excluded"] = round(co9180_subs[sg_name]["excluded"] + excl, 2)
                        break

            # Soaring CO8810 tracking (depts 16 & 17)
            if class_code == "CO8810" and dept_num in SOARING_CO8810_DEPTS:
                soaring_co8810["gross"] = round(soaring_co8810["gross"] + amount, 2)
                soaring_co8810["excluded"] = round(soaring_co8810["excluded"] + excl, 2)

            # Employee-level detail for PDF subgroups
            if dept_num in SUBGROUP_DEPT_NUMS:
                if dept not in subgroup_detail:
                    subgroup_detail[dept] = {}
                if name not in subgroup_detail[dept]:
                    subgroup_detail[dept][name] = {
                        "gross": 0.0, "excluded": 0.0, "class_code": class_code
                    }
                ed = subgroup_detail[dept][name]
                ed["gross"] = round(ed["gross"] + amount, 2)
                ed["excluded"] = round(ed["excluded"] + excl, 2)

    # Build class_codes output
    class_codes = []
    for code in CLASS_CODE_ORDER:
        g = round(totals[code]["gross"], 2)
        ex = round(totals[code]["excluded"], 2)
        class_codes.append({"code": code, "gross": g, "excluded": ex, "compensable": round(g - ex, 2)})

    co9180_subgroups = []
    for sg_name in CO9180_SUBGROUP_DEPTS:
        g = round(co9180_subs[sg_name]["gross"], 2)
        ex = round(co9180_subs[sg_name]["excluded"], 2)
        co9180_subgroups.append({"description": sg_name, "gross": g, "excluded": ex, "compensable": round(g - ex, 2)})

    sc = soaring_co8810
    soaring_co8810_out = {
        "gross": round(sc["gross"], 2),
        "excluded": round(sc["excluded"], 2),
        "compensable": round(sc["gross"] - sc["excluded"], 2),
    }

    total_gross = round(sum(c["gross"] for c in class_codes), 2)
    total_excl = round(sum(c["excluded"] for c in class_codes), 2)

    return {
        "rows_processed": len(payroll_rows),
        "earning_entries": earning_count,
        "class_codes": class_codes,
        "co9180_subgroups": co9180_subgroups,
        "soaring_co8810": soaring_co8810_out,
        "subgroup_detail": subgroup_detail,
        "totals": {
            "gross": total_gross,
            "excluded": total_excl,
            "compensable": round(total_gross - total_excl, 2),
        },
        "overrides_applied": overrides_applied,
        "exceptions": exceptions,
        "dq_notes": dq_notes,
    }


def _build_summary(result: dict, period_date: str) -> str:
    """Build the plain-text summary from pre-calculated classification results."""
    from datetime import datetime
    try:
        dt = datetime.strptime(period_date, "%m/%d/%Y")
        month_label = dt.strftime("%B %Y")
    except Exception:
        month_label = period_date

    overrides = result.get("overrides_applied", [])
    exceptions = result.get("exceptions", [])
    dq_notes = result.get("dq_notes", [])
    totals = result["totals"]

    reassign_count = sum(1 for o in overrides if "Reassign" in o)
    exclude_count = sum(1 for o in overrides if "Exclude" in o)

    lines = [
        f"PAYROLL REPORT SUMMARY — {month_label}",
        f"Reporting period end: {period_date}",
        "",
        "Source file: payroll export",
        f"Detail rows processed: {result['rows_processed']}",
        f"Earning entries processed: {result['earning_entries']}",
        f"Overrides applied: {len(overrides)} ({reassign_count} Reassign, {exclude_count} Exclude)",
        f"Exceptions: {len(exceptions)}",
        "",
        "TOTALS:",
        f"  Gross Payroll       ${totals['gross']:>12,.2f}",
        f"  Excluded Payroll    ${totals['excluded']:>12,.2f}",
        f"  Compensable Payroll ${totals['compensable']:>12,.2f}",
        "",
        "BY CLASS CODE:",
    ]
    for cc in result["class_codes"]:
        lines.append(
            f"  {cc['code']}  Gross ${cc['gross']:,.2f}  "
            f"Excluded ${cc['excluded']:,.2f}  Compensable ${cc['compensable']:,.2f}"
        )

    lines += ["", "OVERRIDES APPLIED:"]
    if overrides:
        for o in overrides:
            lines.append(f"  {o}")
    else:
        lines.append("  (none)")

    lines += ["", "EXCEPTIONS (not included in totals):"]
    if exceptions:
        for e in exceptions:
            lines.append(f"  {e}")
    else:
        lines.append("  (none)")

    if dq_notes:
        lines += ["", "DATA QUALITY NOTES:"]
        for n in dq_notes:
            lines.append(f"  {n}")

    lines += ["", "VARIANCE CHECK:", "  No prior month context available."]
    return "\n".join(lines)


def _fill_form_template(class_codes: list[dict], period_date: str) -> bytes:
    wb = openpyxl.load_workbook(FORM_TEMPLATE_PATH)
    ws = wb.active

    ws["B12"] = period_date

    totals = {"gross": 0.0, "excluded": 0.0, "compensable": 0.0}
    submitted_codes = {item["code"]: item for item in class_codes}

    for code in CLASS_CODE_ORDER:
        row = CODE_TO_ROW[code]
        data = submitted_codes.get(code, {"gross": 0.0, "excluded": 0.0, "compensable": 0.0})
        gross = round(float(data.get("gross", 0)), 2)
        excluded = round(float(data.get("excluded", 0)), 2)
        compensable = round(float(data.get("compensable", 0)), 2)
        ws.cell(row=row, column=6).value = gross
        ws.cell(row=row, column=7).value = excluded
        ws.cell(row=row, column=8).value = compensable
        totals["gross"] += gross
        totals["excluded"] += excluded
        totals["compensable"] += compensable

    ws["F59"] = round(totals["gross"], 2)
    ws["G59"] = round(totals["excluded"], 2)
    ws["H59"] = round(totals["compensable"], 2)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _extract_period_date(text: str) -> str | None:
    # Primary: period tag embedded in the subject line
    tagged = re.search(r"\[period:(\d{1,2}/\d{1,2}/\d{4})\]", text, re.IGNORECASE)
    if tagged:
        return tagged.group(1)
    # Secondary: labelled line
    labelled = re.search(
        r"REPORTING PERIOD END DATE[:\s]+(\d{1,2}/\d{1,2}/\d{4})", text, re.IGNORECASE
    )
    if labelled:
        return labelled.group(1)
    # Last resort: any m/d/yyyy date
    match = re.search(r"\b(\d{1,2}/\d{1,2}/\d{4})\b", text)
    return match.group(1) if match else None


# ── Main pipeline ─────────────────────────────────────────────────────────────

async def process_payroll(xlsx_bytes: bytes, period_date: str) -> tuple[bytes, bytes, bytes, str]:
    """
    Claude classifies each payroll row (class code, overrides, exceptions).
    Python aggregates all dollar arithmetic from those decisions.
    Reports match the Claude Project output because Claude uses the same rules.
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    if "Earnings by Department" not in wb.sheetnames:
        raise ValueError("Sheet 'Earnings by Department' not found in the uploaded file.")

    ws = wb["Earnings by Department"]
    payroll_rows = extract_payroll_rows(ws)
    mapping_rules = load_mapping_rules(MAPPING_RULES_PATH)

    print(f"  Extracted {len(payroll_rows)} payroll rows — sending to Claude for classification...")
    classification = await _classify_with_claude(payroll_rows, period_date)
    print(f"  Claude classified {len(classification.get('classified_rows', []))} rows")

    result = _aggregate(payroll_rows, classification, mapping_rules)

    for exc in result.get("exceptions", []):
        print(f"  EXCEPTION: {exc}")
    for ov in result.get("overrides_applied", []):
        print(f"  OVERRIDE:  {ov}")
    for dq in result.get("dq_notes", []):
        print(f"  DQ NOTE:   {dq}")

    form_bytes = _fill_form_template(result["class_codes"], period_date)
    pdf_bytes = generate_pdf(
        period_date=period_date,
        class_codes=result["class_codes"],
        co9180_subgroups=result.get("co9180_subgroups", []),
        soaring_co8810=result.get("soaring_co8810"),
        subgroup_detail=result.get("subgroup_detail", {}),
    )
    broadmoor_bytes = generate_broadmoor_pdf(
        period_date=period_date,
        subgroup_detail=result.get("subgroup_detail", {}),
        soaring_co8810=result.get("soaring_co8810"),
        co9180_subgroups=result.get("co9180_subgroups", []),
    )
    summary = _build_summary(result, period_date)
    return form_bytes, pdf_bytes, broadmoor_bytes, summary


def _coordinator_emails() -> list[str]:
    raw = os.getenv("COORDINATOR_EMAIL", "")
    return [e.strip() for e in raw.split(",") if e.strip()]


async def handle_inbound(payload: dict):
    """Orchestrate the full pipeline for an inbound coordinator email."""
    coordinators = _coordinator_emails()
    from_email = payload.get("from_email", "").lower()

    if not any(c.lower() in from_email for c in coordinators):
        print(f"Ignoring inbound email from unexpected sender: {from_email}")
        return

    xlsx_bytes = None
    xlsx_filename = None
    for att in payload.get("attachment_files", []):
        if att["filename"].lower().endswith(".xlsx"):
            xlsx_bytes = att["content"]
            xlsx_filename = att["filename"]
            break

    if not xlsx_bytes:
        print("Inbound email from coordinator had no .xlsx attachment — ignoring")
        return

    combined_text = payload.get("subject", "") + " " + payload.get("body", "")
    period_date = _extract_period_date(combined_text)
    if not period_date:
        today = date.today()
        last_day = today.replace(day=1) - timedelta(days=1)
        period_date = f"{last_day.month}/{last_day.day}/{last_day.year}"
        print(f"No period date found in email — defaulting to {period_date}")

    print(f"Processing payroll for period ending {period_date}, file: {xlsx_filename}")

    try:
        form_bytes, pdf_bytes, broadmoor_bytes, summary = await process_payroll(xlsx_bytes, period_date)
        for recipient in coordinators:
            await send_result_email(recipient, period_date, form_bytes, pdf_bytes, broadmoor_bytes, summary)
        print(f"Completed and emailed report for period {period_date} to {coordinators}")
    except Exception as exc:
        import traceback
        error_msg = traceback.format_exc()
        print(f"Error processing payroll:\n{error_msg}")
        for recipient in coordinators:
            await send_error_email(recipient, period_date, str(exc))
